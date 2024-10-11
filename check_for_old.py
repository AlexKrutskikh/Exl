import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Путь к файлу old
old_file_path = r"C:\Users\kruts\PycharmProjects\Exl\exlfile\old_updated.xlsx"

# Чтение данных из файла old
old_data = pd.read_excel(old_file_path)

# Удаление лишних пробелов в заголовках
old_data.columns = old_data.columns.str.strip()

# Загружаем файл для изменения форматирования
wb = load_workbook(old_file_path)
ws = wb.active

# Определяем стиль для текста (красный цвет)
red_font = Font(color="FF0000")

# Цикл по всем четным столбцам
for col in range(2, len(old_data.columns) + 1, 2):  # Начинаем с 2 и проходим по четным столбцам
    for row in range(2, len(old_data) + 2):  # Начинаем с 2, чтобы пропустить заголовок
        odd_value = str(old_data.iloc[row - 2, col - 2])  # Значение из соседнего нечетного столбца
        even_value = str(old_data.iloc[row - 2, col - 1])  # Значение из четного столбца

        if odd_value != even_value:  # Если значения различаются
            # Окрашиваем текст в ячейке четного столбца
            cell = ws.cell(row=row, column=col)
            cell.font = red_font  # Окрашиваем текст в ячейке четного столбца

# Сохраняем изменения в файле old
wb.save(old_file_path)

print("Обновление завершено. Текст в четных ячейках с различиями окрашен в красный цвет.")
